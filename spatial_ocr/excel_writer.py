"""
Excel writer: groups ImageTable objects by column count and writes them to .xlsx files.
Mirrors the logic of the Go excel_build.go — one sheet per column count in a single workbook.
"""
import os
from typing import Dict, List

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from models import ImageTable

MAX_SHEET_NAME_LENGTH = 31


def write_single_workbook(output_dir: str, by_cols: Dict[int, List[ImageTable]]) -> None:
    """
    Writes one .xlsx file with one sheet per column count (e.g. 4_cols, 8_cols).
    Rows from images with the same column count are stacked vertically.
    Duplicate header rows are suppressed across images on the same sheet.
    """
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, "tables_by_column_count.xlsx")

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)  # remove default empty sheet

    for col_count in sorted(by_cols.keys()):
        tables = by_cols[col_count]
        sheet_name = _build_sheet_name(col_count)
        sheet = workbook.create_sheet(title=sheet_name)
        _write_tables_to_sheet(sheet, tables)

    workbook.save(output_path)
    # Intentionally quiet during batch runs.


def write_workbooks_by_column_count(output_dir: str, by_cols: Dict[int, List[ImageTable]]) -> None:
    """
    Writes one .xlsx file per column count (e.g. tables_4_cols.xlsx, tables_8_cols.xlsx).
    """
    os.makedirs(output_dir, exist_ok=True)

    for col_count in sorted(by_cols.keys()):
        tables = by_cols[col_count]
        file_path = os.path.join(output_dir, f"tables_{col_count}_cols.xlsx")
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = _build_sheet_name(col_count)
        _write_tables_to_sheet(sheet, tables)
        workbook.save(file_path)
        # Intentionally quiet during batch runs.


def _write_tables_to_sheet(sheet: Worksheet, tables: List[ImageTable]) -> None:
    """Appends all image table blocks to a worksheet, suppressing duplicate headers."""
    header_row: List[str] = []
    has_header = False

    for index, table in enumerate(tables):
        rows = list(table.rows)

        if index == 0 and rows:
            header_row = rows[0]
            has_header = True

        if index > 0 and has_header and rows and _is_same_row(header_row, rows[0]):
            rows = rows[1:]

        for row_cells in rows:
            sheet.append(row_cells)

        if index < len(tables) - 1:
            sheet.append([])  # blank separator row between image blocks


def _is_same_row(row_a: List[str], row_b: List[str]) -> bool:
    """Returns True if two rows have identical content after stripping whitespace."""
    if len(row_a) != len(row_b):
        return False
    return all(a.strip() == b.strip() for a, b in zip(row_a, row_b))


def _build_sheet_name(col_count: int) -> str:
    name = f"{col_count}_cols"
    return name[:MAX_SHEET_NAME_LENGTH]
